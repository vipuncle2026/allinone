"""车辆管理模块 API"""
import io
import re
from typing import Optional
from datetime import datetime
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import desc, func as sqlfunc
from database import get_db
from models.vehicle import Vehicle, FuelRecord, VehicleExpense

router = APIRouter(prefix="/api/vehicle", tags=["车辆管理"])


def _parse_datetime_str(s):
    """解析日期字符串，兼容 YYYY-MM-DD 和 YYYY-MM-DD HH:mm:ss 格式"""
    if not s:
        return None
    s = s.strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    raise ValueError(f"无法解析日期: {s}")


# ─── Schemas ────────────────────────────────────────────────

class VehicleCreate(BaseModel):
    name: str
    brand: str = ""
    model: str = ""
    color: str = ""
    plate: str = ""
    fuel_type: str = "汽油"
    purchase_date: str = None
    purchase_price: float = None
    current_mileage: float = 0
    notes: str = ""

class VehicleUpdate(BaseModel):
    name: str = None
    brand: str = None
    model: str = None
    color: str = None
    plate: str = None
    fuel_type: str = None
    purchase_date: str = None
    purchase_price: float = None
    current_mileage: float = None
    notes: str = None

class FuelCreate(BaseModel):
    vehicle_id: int
    fuel_date: str
    total_mileage: Optional[float] = None
    unit_price: Optional[float] = None
    fuel_amount: Optional[float] = None
    display_cost: Optional[float] = None
    actual_cost: Optional[float] = None
    fuel_grade: str = ""
    is_full: bool = False
    is_low_fuel: bool = False
    is_missed: bool = False
    fuel_consumption: Optional[float] = None
    station_name: str = ""
    notes: str = ""
    # 电动车字段
    energy_kwh: Optional[float] = None
    electricity_price: Optional[float] = None
    electricity_consumption: Optional[float] = None
    charge_type: str = ""
    soc_start: Optional[int] = None
    soc_end: Optional[int] = None

class FuelUpdate(BaseModel):
    vehicle_id: Optional[int] = None
    fuel_date: Optional[str] = None
    total_mileage: Optional[float] = None
    unit_price: Optional[float] = None
    fuel_amount: Optional[float] = None
    display_cost: Optional[float] = None
    actual_cost: Optional[float] = None
    fuel_grade: Optional[str] = None
    is_full: Optional[bool] = None
    is_low_fuel: Optional[bool] = None
    is_missed: Optional[bool] = None
    fuel_consumption: Optional[float] = None
    station_name: Optional[str] = None
    notes: Optional[str] = None
    # 电动车字段
    energy_kwh: Optional[float] = None
    electricity_price: Optional[float] = None
    electricity_consumption: Optional[float] = None
    charge_type: Optional[str] = None
    soc_start: Optional[int] = None
    soc_end: Optional[int] = None

class ExpenseCreate(BaseModel):
    vehicle_id: int
    expense_date: str = ""
    expense_type: str = ""
    amount: Optional[float] = None
    mileage_at: Optional[float] = None
    notes: str = ""

class ExpenseUpdate(BaseModel):
    vehicle_id: Optional[int] = None
    expense_date: Optional[str] = None
    expense_type: Optional[str] = None
    amount: Optional[float] = None
    mileage_at: Optional[float] = None
    notes: Optional[str] = None


# ─── 车辆 CRUD ─────────────────────────────────────────────

@router.get("/list")
def list_vehicles(db: Session = Depends(get_db)):
    vehicles = db.query(Vehicle).order_by(Vehicle.id).all()
    result = []
    for v in vehicles:
        fuel_count = db.query(FuelRecord).filter(FuelRecord.vehicle_id == v.id).count()
        expense_total = db.query(sqlfunc.coalesce(sqlfunc.sum(VehicleExpense.amount), 0))\
            .filter(VehicleExpense.vehicle_id == v.id).scalar()
        fuel_cost_total = db.query(sqlfunc.coalesce(sqlfunc.sum(FuelRecord.actual_cost), 0))\
            .filter(FuelRecord.vehicle_id == v.id).scalar()
        expense_count = db.query(VehicleExpense).filter(VehicleExpense.vehicle_id == v.id).count()
        total_spending = expense_total + fuel_cost_total
        result.append({
            "id": v.id, "name": v.name, "brand": v.brand, "model": v.model,
            "color": v.color, "plate": v.plate, "fuel_type": v.fuel_type,
            "purchase_date": str(v.purchase_date) if v.purchase_date else None,
            "purchase_price": v.purchase_price, "current_mileage": v.current_mileage,
            "notes": v.notes, "fuel_count": fuel_count, "expense_count": expense_count,
            "total_spending": total_spending,
        })
    return result


@router.post("/create")
def create_vehicle(data: VehicleCreate, db: Session = Depends(get_db)):
    v = Vehicle(**data.model_dump())
    if data.purchase_date:
        v.purchase_date = datetime.strptime(data.purchase_date, "%Y-%m-%d").date()
    db.add(v)
    db.commit()
    db.refresh(v)
    return {"id": v.id, "message": "创建成功"}


@router.put("/{vehicle_id}")
def update_vehicle(vehicle_id: int, data: VehicleUpdate, db: Session = Depends(get_db)):
    v = db.query(Vehicle).filter(Vehicle.id == vehicle_id).first()
    if not v:
        raise HTTPException(404, "车辆不存在")
    update_data = data.model_dump(exclude_none=True)
    if "purchase_date" in update_data and update_data["purchase_date"]:
        update_data["purchase_date"] = datetime.strptime(update_data["purchase_date"], "%Y-%m-%d").date()
    for k, val in update_data.items():
        setattr(v, k, val)
    db.commit()
    return {"message": "更新成功"}


@router.delete("/{vehicle_id}")
def delete_vehicle(vehicle_id: int, db: Session = Depends(get_db)):
    v = db.query(Vehicle).filter(Vehicle.id == vehicle_id).first()
    if not v:
        raise HTTPException(404, "车辆不存在")
    db.query(FuelRecord).filter(FuelRecord.vehicle_id == vehicle_id).delete()
    db.query(VehicleExpense).filter(VehicleExpense.vehicle_id == vehicle_id).delete()
    db.delete(v)
    db.commit()
    return {"message": "删除成功"}


# ─── 油耗记录 CRUD ─────────────────────────────────────────

@router.get("/fuel/list")
def list_fuel(vehicle_id: int, page: int = 1, page_size: int = 20, db: Session = Depends(get_db)):
    total = db.query(FuelRecord).filter(FuelRecord.vehicle_id == vehicle_id).count()
    # 先按日期正序查全部（用于计算行程），再翻转返回
    all_records = db.query(FuelRecord).filter(FuelRecord.vehicle_id == vehicle_id)\
        .order_by(FuelRecord.fuel_date).all()
    # 计算每条记录的行程 = 当前里程 - 上一条里程
    prev_mileage = None
    for r in all_records:
        cur_mileage = r.total_mileage
        r._computed_trip = None
        if cur_mileage is not None and prev_mileage is not None:
            diff = cur_mileage - prev_mileage
            if diff > 0:
                r._computed_trip = round(diff, 1)
        prev_mileage = cur_mileage
    # 按日期倒序排列
    all_records.reverse()
    # 分页
    items = all_records[(page - 1) * page_size: page * page_size]
    return {
        "items": [_fuel_to_dict(r) for r in items],
        "total": total, "page": page, "page_size": page_size,
    }


@router.post("/fuel/create")
def create_fuel(data: FuelCreate, db: Session = Depends(get_db)):
    r = FuelRecord(**data.model_dump())
    r.fuel_date = datetime.strptime(data.fuel_date, "%Y-%m-%d %H:%M:%S")
    db.add(r)
    db.commit()
    db.refresh(r)
    # 更新车辆里程
    if data.total_mileage:
        v = db.query(Vehicle).filter(Vehicle.id == data.vehicle_id).first()
        if v and (v.current_mileage or 0) < data.total_mileage:
            v.current_mileage = data.total_mileage
            db.commit()
    return {"id": r.id, "message": "添加成功"}


@router.put("/fuel/{fuel_id}")
def update_fuel(fuel_id: int, data: FuelUpdate, db: Session = Depends(get_db)):
    r = db.query(FuelRecord).filter(FuelRecord.id == fuel_id).first()
    if not r:
        raise HTTPException(404, "记录不存在")
    update_data = data.model_dump(exclude_none=True)
    if "fuel_date" in update_data and update_data["fuel_date"]:
        update_data["fuel_date"] = datetime.strptime(update_data["fuel_date"], "%Y-%m-%d %H:%M:%S")
    for k, val in update_data.items():
        setattr(r, k, val)
    db.commit()
    return {"message": "更新成功"}


@router.delete("/fuel/{fuel_id}")
def delete_fuel(fuel_id: int, db: Session = Depends(get_db)):
    r = db.query(FuelRecord).filter(FuelRecord.id == fuel_id).first()
    if not r:
        raise HTTPException(404, "记录不存在")
    db.delete(r)
    db.commit()
    return {"message": "删除成功"}


# ─── 费用记录 CRUD ─────────────────────────────────────────

@router.get("/expense/list")
def list_expense(vehicle_id: int, page: int = 1, page_size: int = 20, db: Session = Depends(get_db)):
    total = db.query(VehicleExpense).filter(VehicleExpense.vehicle_id == vehicle_id).count()
    records = db.query(VehicleExpense).filter(VehicleExpense.vehicle_id == vehicle_id)\
        .order_by(desc(VehicleExpense.expense_date))\
        .offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": [_expense_to_dict(r) for r in records],
        "total": total, "page": page, "page_size": page_size,
    }


@router.post("/expense/create")
def create_expense(data: ExpenseCreate, db: Session = Depends(get_db)):
    r = VehicleExpense(**data.model_dump())
    if data.expense_date:
        r.expense_date = _parse_datetime_str(data.expense_date)
    else:
        r.expense_date = None
    db.add(r)
    db.commit()
    db.refresh(r)
    return {"id": r.id, "message": "添加成功"}


@router.put("/expense/{expense_id}")
def update_expense(expense_id: int, data: ExpenseUpdate, db: Session = Depends(get_db)):
    r = db.query(VehicleExpense).filter(VehicleExpense.id == expense_id).first()
    if not r:
        raise HTTPException(404, "记录不存在")
    update_data = data.model_dump(exclude_none=True)
    if "expense_date" in update_data:
        update_data["expense_date"] = _parse_datetime_str(update_data["expense_date"]) if update_data["expense_date"] else None
    for k, val in update_data.items():
        setattr(r, k, val)
    db.commit()
    return {"message": "更新成功"}


@router.delete("/expense/{expense_id}")
def delete_expense(expense_id: int, db: Session = Depends(get_db)):
    r = db.query(VehicleExpense).filter(VehicleExpense.id == expense_id).first()
    if not r:
        raise HTTPException(404, "记录不存在")
    db.delete(r)
    db.commit()
    return {"message": "删除成功"}


# ─── 统计 ──────────────────────────────────────────────────

@router.get("/stats")
def vehicle_stats(vehicle_id: int, db: Session = Depends(get_db)):
    fuel_count = db.query(FuelRecord).filter(FuelRecord.vehicle_id == vehicle_id).count()
    total_fuel = db.query(sqlfunc.coalesce(sqlfunc.sum(FuelRecord.fuel_amount), 0))\
        .filter(FuelRecord.vehicle_id == vehicle_id).scalar()
    total_fuel_cost = db.query(sqlfunc.coalesce(sqlfunc.sum(FuelRecord.actual_cost), 0))\
        .filter(FuelRecord.vehicle_id == vehicle_id).scalar()
    avg_consumption = db.query(sqlfunc.avg(FuelRecord.fuel_consumption))\
        .filter(FuelRecord.vehicle_id == vehicle_id, FuelRecord.fuel_consumption > 0).scalar()

    # 电动车统计
    total_energy = db.query(sqlfunc.coalesce(sqlfunc.sum(FuelRecord.energy_kwh), 0))\
        .filter(FuelRecord.vehicle_id == vehicle_id).scalar()
    avg_elec_consumption = db.query(sqlfunc.avg(FuelRecord.electricity_consumption))\
        .filter(FuelRecord.vehicle_id == vehicle_id, FuelRecord.electricity_consumption > 0).scalar()

    expense_count = db.query(VehicleExpense).filter(VehicleExpense.vehicle_id == vehicle_id).count()
    total_expense = db.query(sqlfunc.coalesce(sqlfunc.sum(VehicleExpense.amount), 0))\
        .filter(VehicleExpense.vehicle_id == vehicle_id).scalar()

    # 费用分类统计
    expense_by_type = db.query(VehicleExpense.expense_type, sqlfunc.coalesce(sqlfunc.sum(VehicleExpense.amount), 0))\
        .filter(VehicleExpense.vehicle_id == vehicle_id)\
        .group_by(VehicleExpense.expense_type).all()

    return {
        "fuel_count": fuel_count,
        "total_fuel_liters": round(total_fuel, 2),
        "total_fuel_cost": round(total_fuel_cost, 2),
        "avg_consumption": round(avg_consumption, 2) if avg_consumption else None,
        # 电动车
        "total_energy_kwh": round(total_energy, 2),
        "avg_electricity_consumption": round(avg_elec_consumption, 2) if avg_elec_consumption else None,
        "expense_count": expense_count,
        "total_expense": round(total_expense, 2),
        "total_spending": round(total_fuel_cost + total_expense, 2),
        "expense_by_type": {t: round(a, 2) for t, a in expense_by_type},
    }


@router.get("/stats/analysis")
def vehicle_analysis(vehicle_id: int, db: Session = Depends(get_db)):
    """车辆统计分析：按年、按月的油耗和费用统计"""

    # ─── 按年统计 ────────────────────────────────
    yearly_fuel = db.query(
        sqlfunc.strftime("%Y", FuelRecord.fuel_date).label("period"),
        sqlfunc.count(FuelRecord.id).label("fuel_count"),
        sqlfunc.coalesce(sqlfunc.sum(FuelRecord.fuel_amount), 0).label("total_fuel_liters"),
        sqlfunc.coalesce(sqlfunc.sum(FuelRecord.actual_cost), 0).label("total_fuel_cost"),
        sqlfunc.avg(FuelRecord.fuel_consumption).label("avg_consumption"),
        sqlfunc.min(FuelRecord.unit_price).label("min_price"),
        sqlfunc.max(FuelRecord.unit_price).label("max_price"),
        # 电动车字段
        sqlfunc.coalesce(sqlfunc.sum(FuelRecord.energy_kwh), 0).label("total_energy_kwh"),
        sqlfunc.avg(FuelRecord.electricity_consumption).label("avg_elec_consumption"),
        sqlfunc.min(FuelRecord.electricity_price).label("min_elec_price"),
        sqlfunc.max(FuelRecord.electricity_price).label("max_elec_price"),
    ).filter(
        FuelRecord.vehicle_id == vehicle_id,
        FuelRecord.fuel_date.isnot(None),
    ).group_by(sqlfunc.strftime("%Y", FuelRecord.fuel_date)).all()

    yearly_expense = db.query(
        sqlfunc.strftime("%Y", VehicleExpense.expense_date).label("period"),
        sqlfunc.count(VehicleExpense.id).label("expense_count"),
        sqlfunc.coalesce(sqlfunc.sum(VehicleExpense.amount), 0).label("total_expense"),
    ).filter(
        VehicleExpense.vehicle_id == vehicle_id,
        VehicleExpense.expense_date.isnot(None),
    ).group_by(sqlfunc.strftime("%Y", VehicleExpense.expense_date)).all()

    # ─── 按月统计 ────────────────────────────────
    monthly_fuel = db.query(
        sqlfunc.strftime("%Y-%m", FuelRecord.fuel_date).label("period"),
        sqlfunc.count(FuelRecord.id).label("fuel_count"),
        sqlfunc.coalesce(sqlfunc.sum(FuelRecord.fuel_amount), 0).label("total_fuel_liters"),
        sqlfunc.coalesce(sqlfunc.sum(FuelRecord.actual_cost), 0).label("total_fuel_cost"),
        sqlfunc.avg(FuelRecord.fuel_consumption).label("avg_consumption"),
        sqlfunc.min(FuelRecord.unit_price).label("min_price"),
        sqlfunc.max(FuelRecord.unit_price).label("max_price"),
        # 电动车字段
        sqlfunc.coalesce(sqlfunc.sum(FuelRecord.energy_kwh), 0).label("total_energy_kwh"),
        sqlfunc.avg(FuelRecord.electricity_consumption).label("avg_elec_consumption"),
        sqlfunc.min(FuelRecord.electricity_price).label("min_elec_price"),
        sqlfunc.max(FuelRecord.electricity_price).label("max_elec_price"),
    ).filter(
        FuelRecord.vehicle_id == vehicle_id,
        FuelRecord.fuel_date.isnot(None),
    ).group_by(sqlfunc.strftime("%Y-%m", FuelRecord.fuel_date)).all()

    monthly_expense = db.query(
        sqlfunc.strftime("%Y-%m", VehicleExpense.expense_date).label("period"),
        sqlfunc.count(VehicleExpense.id).label("expense_count"),
        sqlfunc.coalesce(sqlfunc.sum(VehicleExpense.amount), 0).label("total_expense"),
    ).filter(
        VehicleExpense.vehicle_id == vehicle_id,
        VehicleExpense.expense_date.isnot(None),
    ).group_by(sqlfunc.strftime("%Y-%m", VehicleExpense.expense_date)).all()

    # ─── 费用分类统计（全部） ────────────────────
    expense_by_type = db.query(
        VehicleExpense.expense_type,
        sqlfunc.coalesce(sqlfunc.sum(VehicleExpense.amount), 0).label("total"),
        sqlfunc.count(VehicleExpense.id).label("count"),
    ).filter(
        VehicleExpense.vehicle_id == vehicle_id,
        VehicleExpense.expense_date.isnot(None),
    ).group_by(VehicleExpense.expense_type).all()

    # ─── 每年费用分类明细 ────────────────────────
    yearly_expense_by_type = db.query(
        sqlfunc.strftime("%Y", VehicleExpense.expense_date).label("period"),
        VehicleExpense.expense_type,
        sqlfunc.coalesce(sqlfunc.sum(VehicleExpense.amount), 0).label("total"),
    ).filter(
        VehicleExpense.vehicle_id == vehicle_id,
        VehicleExpense.expense_date.isnot(None),
    ).group_by(
        sqlfunc.strftime("%Y", VehicleExpense.expense_date),
        VehicleExpense.expense_type,
    ).all()

    # ─── 汇总函数 ────────────────────────────────
    def build_summary(fuel_list, expense_list, prefix=""):
        """合并油耗和费用数据"""
        exp_map = {r.period: r for r in expense_list}
        result = []
        used_exp_periods = set()

        for r in fuel_list:
            e = exp_map.get(r.period)
            if e:
                used_exp_periods.add(r.period)
            fuel_cost = r.total_fuel_cost or 0
            exp_total = e.total_expense if e else 0
            result.append({
                "period": r.period,
                "fuel_count": r.fuel_count,
                "total_fuel_liters": round(r.total_fuel_liters, 2),
                "total_fuel_cost": round(fuel_cost, 2),
                "avg_consumption": round(r.avg_consumption, 2) if r.avg_consumption else None,
                "min_price": round(r.min_price, 2) if r.min_price else None,
                "max_price": round(r.max_price, 2) if r.max_price else None,
                # 电动车
                "total_energy_kwh": round(r.total_energy_kwh, 2) if hasattr(r, 'total_energy_kwh') else 0,
                "avg_elec_consumption": round(r.avg_elec_consumption, 2) if hasattr(r, 'avg_elec_consumption') and r.avg_elec_consumption else None,
                "min_elec_price": round(r.min_elec_price, 2) if hasattr(r, 'min_elec_price') and r.min_elec_price else None,
                "max_elec_price": round(r.max_elec_price, 2) if hasattr(r, 'max_elec_price') and r.max_elec_price else None,
                "expense_count": e.expense_count if e else 0,
                "total_expense": round(exp_total, 2),
                "total_spending": round(fuel_cost + exp_total, 2),
            })

        # 补充只有费用没有加油的时段
        for period, e in exp_map.items():
            if period not in used_exp_periods:
                exp_total = e.total_expense or 0
                result.append({
                    "period": period,
                    "fuel_count": 0,
                    "total_fuel_liters": 0,
                    "total_fuel_cost": 0,
                    "avg_consumption": None,
                    "min_price": None,
                    "max_price": None,
                    "total_energy_kwh": 0,
                    "avg_elec_consumption": None,
                    "min_elec_price": None,
                    "max_elec_price": None,
                    "expense_count": e.expense_count,
                    "total_expense": round(exp_total, 2),
                    "total_spending": round(exp_total, 2),
                })

        result.sort(key=lambda x: x["period"])
        return result

    yearly_data = build_summary(yearly_fuel, yearly_expense)
    monthly_data = build_summary(monthly_fuel, monthly_expense)

    return {
        "yearly": yearly_data,
        "monthly": monthly_data,
        "expense_by_type": [{"type": r.expense_type or "其他", "total": round(r.total, 2), "count": r.count} for r in expense_by_type],
        "yearly_expense_by_type": [{"period": r.period, "type": r.expense_type or "其他", "total": round(r.total, 2)} for r in yearly_expense_by_type],
    }


# ─── Excel 导入导出 ───────────────────────────────────────

@router.post("/import/xlsx")
async def import_xlsx(vehicle_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """导入小熊油耗 Excel 文件"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "服务端未安装 openpyxl，请执行 pip install openpyxl")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)

    imported = {"fuel": 0, "expense": 0}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if "油耗" in sheet_name:
            # 油耗记录 sheet：Row1=表头
            for row_idx in range(2, ws.max_row + 1):
                vals = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
                if not vals[0]:
                    continue
                try:
                    fuel_date = _parse_datetime(vals[0])
                    if not fuel_date:
                        continue
                    # 检查是否已存在（同一车辆同一时间）
                    exists = db.query(FuelRecord).filter(
                        FuelRecord.vehicle_id == vehicle_id,
                        FuelRecord.fuel_date == fuel_date
                    ).first()
                    if exists:
                        continue

                    record = FuelRecord(vehicle_id=vehicle_id, fuel_date=fuel_date)
                    record.total_mileage = _to_float(vals[1])
                    record.unit_price = _to_float(vals[2])
                    record.fuel_amount = _to_float(vals[3])
                    record.display_cost = _to_float(vals[4])
                    record.actual_cost = _to_float(vals[5])
                    record.fuel_grade = str(vals[6] or "")
                    record.is_full = vals[7] == "是" if vals[7] else False
                    record.is_low_fuel = vals[8] == "是" if vals[8] else False
                    record.is_missed = vals[9] == "是" if vals[9] else False
                    record.fuel_consumption = _to_float(vals[10])
                    record.trip_mileage = _to_float(vals[11])
                    record.station_name = str(vals[12] or "")
                    record.notes = str(vals[13] or "") if len(vals) > 13 and vals[13] else ""
                    db.add(record)
                    imported["fuel"] += 1

                    # 更新车辆里程
                    if record.total_mileage:
                        v = db.query(Vehicle).filter(Vehicle.id == vehicle_id).first()
                        if v and (v.current_mileage or 0) < record.total_mileage:
                            v.current_mileage = record.total_mileage
                except Exception as e:
                    print(f"导入油耗行 {row_idx} 失败: {e}")
                    continue

        elif "费用" in sheet_name:
            # 费用记录 sheet
            for row_idx in range(2, ws.max_row + 1):
                vals = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
                if not vals[0]:
                    continue
                try:
                    expense_date = _parse_datetime(vals[0])
                    if not expense_date:
                        continue
                    exists = db.query(VehicleExpense).filter(
                        VehicleExpense.vehicle_id == vehicle_id,
                        VehicleExpense.expense_date == expense_date,
                        VehicleExpense.expense_type == (vals[1] or "")
                    ).first()
                    if exists:
                        continue

                    record = VehicleExpense(vehicle_id=vehicle_id, expense_date=expense_date)
                    record.expense_type = str(vals[1] or "")
                    record.amount = _to_float(vals[2])
                    record.notes = str(vals[3] or "") if len(vals) > 3 and vals[3] else ""
                    db.add(record)
                    imported["expense"] += 1
                except Exception as e:
                    print(f"导入费用行 {row_idx} 失败: {e}")
                    continue

    db.commit()
    return {"imported": imported, "message": f"导入完成：{imported['fuel']}条油耗，{imported['expense']}条费用"}


@router.get("/export/xlsx")
def export_xlsx(vehicle_id: int, db: Session = Depends(get_db)):
    """导出车辆数据为 Excel（小熊油耗兼容格式）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        raise HTTPException(500, "服务端未安装 openpyxl")

    wb = Workbook()

    # Sheet 1: 油耗记录
    ws_fuel = wb.active
    ws_fuel.title = "油耗记录"
    headers = ["日期时间", "总里程", "机显单价", "加油量", "机显金额", "实付金额",
               "油号", "加满", "亮灯", "漏记", "油耗", "行程", "加油站名称", "备注"]
    for c, h in enumerate(headers, 1):
        cell = ws_fuel.cell(1, c, h)
        cell.font = Font(bold=True)

    records = db.query(FuelRecord).filter(FuelRecord.vehicle_id == vehicle_id)\
        .order_by(FuelRecord.fuel_date).all()
    for r_idx, r in enumerate(records, 2):
        ws_fuel.cell(r_idx, 1, r.fuel_date.strftime("%Y-%m-%d %H:%M:%S") if r.fuel_date else "")
        ws_fuel.cell(r_idx, 2, r.total_mileage)
        ws_fuel.cell(r_idx, 3, r.unit_price)
        ws_fuel.cell(r_idx, 4, r.fuel_amount)
        ws_fuel.cell(r_idx, 5, r.display_cost)
        ws_fuel.cell(r_idx, 6, r.actual_cost)
        ws_fuel.cell(r_idx, 7, r.fuel_grade)
        ws_fuel.cell(r_idx, 8, "是" if r.is_full else "否")
        ws_fuel.cell(r_idx, 9, "是" if r.is_low_fuel else "否")
        ws_fuel.cell(r_idx, 10, "是" if r.is_missed else "否")
        ws_fuel.cell(r_idx, 11, r.fuel_consumption)
        ws_fuel.cell(r_idx, 12, r.trip_mileage)
        ws_fuel.cell(r_idx, 13, r.station_name)
        ws_fuel.cell(r_idx, 14, r.notes)

    for col in range(1, 15):
        ws_fuel.column_dimensions[ws_fuel.cell(1, col).column_letter].width = 16

    # Sheet 2: 费用记录
    ws_exp = wb.create_sheet("费用记录")
    exp_headers = ["日期时间", "费用类型名称", "金额", "备注"]
    for c, h in enumerate(exp_headers, 1):
        cell = ws_exp.cell(1, c, h)
        cell.font = Font(bold=True)

    v = db.query(Vehicle).filter(Vehicle.id == vehicle_id).first()
    vname = v.name if v else "未知"
    ws_exp.title = f"{vname}-费用记录"

    expenses = db.query(VehicleExpense).filter(VehicleExpense.vehicle_id == vehicle_id)\
        .order_by(VehicleExpense.expense_date).all()
    for r_idx, e in enumerate(expenses, 2):
        ws_exp.cell(r_idx, 1, e.expense_date.strftime("%Y-%m-%d %H:%M:%S") if e.expense_date else "")
        ws_exp.cell(r_idx, 2, e.expense_type)
        ws_exp.cell(r_idx, 3, e.amount)
        ws_exp.cell(r_idx, 4, e.notes)

    for col in range(1, 5):
        ws_exp.column_dimensions[ws_exp.cell(1, col).column_letter].width = 20

    # 输出
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"vehicle_{vehicle_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ─── 工具函数 ──────────────────────────────────────────────

def _fuel_to_dict(r):
    """将油耗记录转为字典，行程由 list_fuel 中预计算的 _computed_trip 提供"""
    return {
        "id": r.id, "vehicle_id": r.vehicle_id,
        "fuel_date": r.fuel_date.strftime("%Y-%m-%d %H:%M:%S") if r.fuel_date else None,
        "total_mileage": r.total_mileage, "unit_price": r.unit_price,
        "fuel_amount": r.fuel_amount, "display_cost": r.display_cost,
        "actual_cost": r.actual_cost, "fuel_grade": r.fuel_grade,
        "is_full": r.is_full, "is_low_fuel": r.is_low_fuel, "is_missed": r.is_missed,
        "fuel_consumption": r.fuel_consumption,
        "trip_mileage": getattr(r, '_computed_trip', None),
        "station_name": r.station_name, "notes": r.notes,
        # 电动车字段
        "energy_kwh": r.energy_kwh,
        "electricity_price": r.electricity_price,
        "electricity_consumption": r.electricity_consumption,
        "charge_type": r.charge_type,
        "soc_start": r.soc_start,
        "soc_end": r.soc_end,
    }

def _expense_to_dict(r):
    return {
        "id": r.id, "vehicle_id": r.vehicle_id,
        "expense_date": r.expense_date.strftime("%Y-%m-%d %H:%M:%S") if r.expense_date else None,
        "expense_type": r.expense_type, "amount": r.amount,
        "mileage_at": r.mileage_at, "notes": r.notes,
    }

def _parse_datetime(val):
    """解析日期时间值"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        s = val.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
    return None

def _to_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


# ─── 最近记录（跨车辆合并，用于 Dashboard） ─────────────

@router.get("/recent")
def get_recent_records(limit: int = 10, db: Session = Depends(get_db)):
    """获取跨车辆的最近加油+费用记录"""
    # 获取所有车辆名称映射
    vehicles = db.query(Vehicle.id, Vehicle.name).all()
    v_map = {v.id: v.name for v in vehicles}

    # 最近加油记录
    fuels = db.query(FuelRecord).order_by(desc(FuelRecord.fuel_date)).limit(limit).all()
    fuel_items = []
    for f in fuels:
        fuel_items.append({
            "record_type": "fuel",
            "id": f.id,
            "date": f.fuel_date.strftime("%Y-%m-%d") if f.fuel_date else "",
            "vehicle": v_map.get(f.vehicle_id, "未知"),
            "title": f"加油 {f.fuel_amount or 0:.1f}L" if f.fuel_amount else "加油",
            "amount": f.actual_cost or f.display_cost or 0,
            "mileage": f.total_mileage,
        })

    # 最近费用记录
    expenses = db.query(VehicleExpense).order_by(desc(VehicleExpense.expense_date)).limit(limit).all()
    exp_items = []
    for e in expenses:
        exp_items.append({
            "record_type": "expense",
            "id": e.id,
            "date": e.expense_date.strftime("%Y-%m-%d") if e.expense_date else "",
            "vehicle": v_map.get(e.vehicle_id, "未知"),
            "title": e.expense_type or "费用",
            "amount": e.amount or 0,
            "mileage": e.mileage_at,
        })

    # 合并按日期排序
    all_items = fuel_items + exp_items
    all_items.sort(key=lambda x: x["date"], reverse=True)
    return all_items[:limit]
